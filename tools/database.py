import datetime
import os
from functools import reduce
from pathlib import Path
from typing import Tuple, Union
from zipfile import ZipFile

from openpyxl import load_workbook

from exceptions import API_FILE, ApiTypeError, TokenError
from models.db_model import Address, City, Map, Region

from models.singleton import Singleton
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from yadisk import YaDisk, exceptions
from helpers import read_config


class Sheet(metaclass=Singleton):
    """
    Class for working with the Google/Yandex tables
    It is used to receive and send the necessary information in the table
    """

    def __yandex_connect(self):
        # check token in config file
        config = read_config()
        token = dict(config).get("YANDEX_API").get("oauth_token")
        if token is None:
            raise TokenError("Can't read oauth_token from config file")

        # check if the token is valid
        disk = YaDisk(token=token)
        if not disk.check_token():
            raise TokenError("Can't connect to Yandex Disk API with your oauth_token")

        self._connection = True
        self._obj = "YANDEX"

        # downloading file
        # check if file for download is specified
        wanted_file = dict(config).get("YANDEX_API").get("wanted_file")
        if wanted_file is None:
            raise API_FILE
        try:
            disk.download(
                wanted_file,
                os.path.dirname(__file__)
                + dict(config).get("COMMON").get("sheet_file"),
            )
        except exceptions.PathNotFoundError:
            print("Not found requested file in yandex disk")
            os.remove(
                os.path.dirname(__file__) + dict(config).get("COMMON").get("sheet_file")
            )

    def __google_connect(self):
        print("Deprecated because of google's policy. Please use Yandex instead.")

    _connection = None
    _obj = None
    _supported_api = {
        "google": {
            "short": ["g"],
            "connection": __google_connect,
        },
        "yandex": {
            "short": ["y"],
            "connection": __yandex_connect,
        },
    }

    @classmethod
    def connect(cls, api_type: str):
        """
        Check connection and download base

        :type argument for Google or Yandex API
        available types - (g)oogle | (y)andex
        """
        if cls._connection is not None:
            return
        api_type = api_type.lower()
        if api_type not in set(cls._supported_api.keys()) | set(
            reduce(
                lambda a, b: a + b,
                map(lambda v: v["short"], cls._supported_api.values()),
            )
        ):
            raise ApiTypeError
        connection_func = None
        if cls._supported_api.get(api_type):
            connection_func = cls._supported_api[api_type]["connection"]
        else:
            for v in cls._supported_api.values():
                if api_type in v["short"]:
                    connection_func = v["connection"]
                    break
        connection_func(cls)

    @classmethod
    def get_connection(cls):
        return cls._obj


class SheetParser(metaclass=Singleton):
    _file = None

    def __init__(cls, file: str = None) -> None:
        super().__init__()
        # if cls._file is None:
        #     if file is None:
        #         config = read_config()
        #         file = dict(config).get("COMMON").get("sheet_file")
        #     cls._file = file
        if cls._file is None:
            cls._file = file

    @classmethod
    def set_file(cls, file: str = None) -> None:
        cls._file = file

    @classmethod
    def get_file(cls):
        return cls._file


class Database(metaclass=Singleton):
    _session = None

    @classmethod
    def init_session(cls, engine=None, echo=True) -> bool:
        if cls._session is None:
            if engine is None or engine in ["sql"]:
                engine = create_engine(
                    "mysql+pymysql://root:pingus@localhost/bdsm", echo=echo
                )
            elif engine in ["sqlite"]:
                engine = create_engine("sqlite:///data/base.db", echo=True)
            cls._session = sessionmaker(bind=engine)()
        return bool(cls._session)

    @classmethod
    def get_map(cls, map_name: str, city: str) -> Union[Map, None]:
        if cls.init_session():
            return (
                cls._session.query(Map)
                .join(Region, Map.region_id == Region.region_id)
                .join(City, Region.city_id == City.city_id)
                .filter(City.name.like(city))
                .filter(Map.name.like(map_name))
                .first()
            )

    @classmethod
    def get_cities(cls) -> Union[list[City], None]:
        if cls.init_session():
            return cls._session.query(City).all()

    @classmethod
    def get_addresses(cls, city=None) -> Union[list[Address], None]:
        if cls.init_session():
            if city is None:
                return cls._session.query(Address).all()
            else:
                return (
                    cls._session.query(Address)
                    .join(Map, Address.map_id == Map.map_id)
                    .join(Region, Region.region_id == Map.region_id)
                    .join(City, City.city_id == Region.city_id)
                    .filter(City.name.like(city))
                    .all()
                )

    @classmethod
    def get_maps(cls, city: str = None) -> Union[list[Map], None]:
        if cls.init_session():
            if city is None:
                return cls._session.query(Map).all()
            return (
                cls._session.query(Map)
                .join(Region, Map.region_id == Region.region_id)
                .join(City, Region.city_id == City.city_id)
                .filter(City.name.like(city))
                .all()
            )

    @classmethod
    def get_matches_map(cls, address: Tuple[str, str], city) -> Union[Map, None]:
        if cls.init_session():
            return (
                cls._session.query(Address)
                .join(Map, Map.map_id == Address.map_id)
                .join(Region, Region.region_id == Map.region_id)
                .join(City, City.city_id == Region.city_id)
                .filter(City.name.like(city))
                .filter(Address.street.like(address[0]))
                .filter(Address.number.like(address[1]))
                .first()
                .map
            )

    @classmethod
    def _collect_maps_zip_xlsx(cls, zip_filepath, xlsx_filepath, city_name):
        def unpack_zip(filepath, city_name):
            maps = {city_name: {}}
            with ZipFile(filepath) as archive:
                for file in filter(
                    lambda filename: filename.endswith(".png"), archive.namelist()
                ):
                    region = Path(file).parent.name
                    if region not in maps.get(city_name):
                        maps[city_name][region] = {}
                    maps[city_name][region][Path(file).stem] = {
                        "picture": archive.read(file),
                        # "picture": "FUCKING BYTES",
                        "addresses": [],
                    }
            return maps

        maps = unpack_zip(zip_filepath, city_name)
        excel_table = load_workbook(xlsx_filepath)

        for city_regions in maps.values():
            for region, region_maps in city_regions.items():
                for map_name, map_values in region_maps.items():

                    map_addresses = list(
                        map(
                            lambda row: list(map(lambda cell: cell.value, row))[1:8],
                            filter(
                                lambda row: row[0].value
                                and row[0].value.lower() == map_name.lower(),
                                excel_table[region].iter_rows(),
                            ),
                        )
                    )
                    for i in range(len(map_addresses.copy())):
                        try:
                            map_addresses[i] = [
                                *map_addresses[i][:5],
                                map_addresses[i][6],
                            ]
                        except IndexError:
                            map_addresses[i] = [*map_addresses[i][:5], None]
                        for j in range(len(map_addresses[i])):
                            if element := map_addresses[i][j]:
                                if isinstance(element, datetime.datetime):
                                    map_addresses[i][
                                        j
                                    ] = f"{element.day}/{element.month}"
                                if str(element).strip() == "-":
                                    map_addresses[i][j] = None
                                else:
                                    map_addresses[i][j] = str(map_addresses[i][j])

                    map_values["addresses"] = map_addresses
        return maps

    @classmethod
    def create_city(cls, city, zip_filepath, xlsx_filepath):
        if not cls.init_session():
            return
        maps = cls._collect_maps_zip_xlsx(zip_filepath, xlsx_filepath, city)

        for city, city_regions in maps.items():
            cls._session.add(City(city))
            for city_region, region_maps in city_regions.items():
                cls._session.add(
                    Region(
                        city_region,
                        cls._session.query(City)
                        .filter_by(name=city)
                        .first()
                        .city_id,
                    ),
                )
                for region_map_number, region_map_vals in region_maps.items():
                    cls._session.add(
                        Map(
                            region_map_number,
                            region_map_vals.get("picture"),
                            cls._session.query(Region)
                            .filter_by(
                                name=city_region,
                                city_id=cls._session.query(City)
                                .filter_by(name=city)
                                .first()
                                .city_id,
                            )
                            .first()
                            .region_id,
                        )
                    )
                    for address in region_map_vals.get("addresses"):
                        cls._session.add(
                            Address(
                                address[0],
                                address[1],
                                address[2],
                                address[3],
                                address[4],
                                cls._session.query(Map)
                                .filter_by(
                                    name=region_map_number,
                                    region_id=cls._session.query(Region)
                                    .filter_by(
                                        name=city_region,
                                        city_id=cls._session.query(City)
                                        .filter_by(name=city)
                                        .first()
                                        .city_id,
                                    )
                                    .first()
                                    .region_id,
                                )
                                .first()
                                .map_id,
                                None,
                                address[5],
                            )
                        )

        cls._session.commit()


# Database.create_city(
#     "Нижний Тагил",
#     "/home/god/Documents/bdsm/bdsm_promo/tools/Карты.zip",
#     "/home/god/Documents/bdsm/bdsm_promo/tools/База Нижний Тагил.xlsx",
# )
