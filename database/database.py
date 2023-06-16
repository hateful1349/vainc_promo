import datetime
from pathlib import Path
from typing import Optional
from zipfile import ZipFile

from openpyxl import load_workbook
from sqlalchemy import func

from specials.singleton import Singleton

from .base import db_session
from .models import Address, City, Map, Region


class Database(metaclass=Singleton):
    """
    A database object that can be used as a context manager
    """

    @classmethod
    def get_city(cls, city_name=None, city_id=None) -> City | None:
        """
        Gets a city by id or name. DON'T USE city_name and city_id TOGETHER

        :param city_name: Name of the city
        :type city_name: str
        :param city_id: ID of the city
        :type city_id: int
        :return: City object
        """

        if city_name and city_id:
            print("DON'T USE city_name and city_id TOGETHER")
            raise Exception
        if not city_name and not city_id:
            print("Empty search")
            raise Exception
        with db_session() as session:
            if city_name:
                return session.query(City).filter(City.name.ilike(city_name)).first()
            return session.query(City).filter(City.id == city_id).first()

    @classmethod
    def get_map(cls, city, map_name=None, map_id=None) -> Map | None:
        """
        Gets a map from a given city

        :param city: The name of the city
        :type city: str
        :param map_name: The name of the map
        :type map_name: str
        :param map_id: Replacement of map_name
        :type map_id: int
        :return: The corresponding map object or None if no map is found
        """
        if not map_name and not map_id:
            print("Empty search")
            raise Exception
        if map_name and map_id:
            print("DONT USE map_name AND map_id TOGETHER")
            raise Exception
        with db_session() as session:
            if map_name:
                return (
                    session.query(Map)
                    .join(Region, Map.region_id == Region.id)
                    .join(City, Region.city_id == City.id)
                    .filter(City.name.ilike(city))
                    .filter(Map.name.ilike(map_name))
                    .first()
                )
            return (
                session.query(Map)
                .join(Region, Map.region_id == Region.id)
                .join(City, Region.city_id == City.id)
                .filter(City.name.ilike(city))
                .filter(Map.id == map_id)
                .first()
            )

    @classmethod
    def get_cities(cls) -> list[City]:
        """
        Gets a list of cities

        :return: The list of cities or None if no cities were found
        """
        with db_session() as session:
            return session.query(City).all()

    @classmethod
    def get_regions(cls, city_name) -> list[Region]:
        """
        Gets a list of regions

        :param city_name: The name of the city
        :type city_name: str
        :return: The list of regions or None if no regions were found
        """
        with db_session() as session:
            return (
                session.query(Region)
                .join(City, City.id == Region.city_id)
                .filter(City.name.ilike(city_name))
                .all()
            )

    @classmethod
    def get_addresses(cls, city_name=None, map_name=None, map_id=None) -> list[Address]:
        """
        Gets a list of addresses

        :param city_name: The name of the city
        :type city_name: str
        :param map_name: The name of the map (use with city_name)
        :type map_name: str
        :param map_id: ID of the map (use with empty city_name and map_name)
        :type map_id: int
        :return: The list of addresses
        """
        if (not city_name or map_id) and (city_name or map_name):
            print("Search by: empty | city_name | city_name+map_name | map_id")
            raise Exception

        with db_session() as session:
            if not any([city_name, map_name, map_id]):
                return session.query(Address).all()
            if city_name and map_name:
                return (
                    session.query(Address)
                    .join(Map, Address.map_id == Map.id)
                    .join(Region, Region.id == Map.region_id)
                    .join(City, City.id == Region.city_id)
                    .filter(City.name.ilike(city_name))
                    .filter(Map.name.ilike(map_name))
                    .all()
                )
            if city_name:
                return (
                    session.query(Address)
                    .join(Map, Address.map_id == Map.id)
                    .join(Region, Region.id == Map.region_id)
                    .join(City, City.id == Region.city_id)
                    .filter(City.name.ilike(city_name))
                    .all()
                )
            return session.query(Address).filter(Address.map_id == map_id).all()

    @classmethod
    def get_maps(cls, city=None, region=None) -> list[Map]:
        """
        Gets a list of maps

        :param city: The name of the city
        :type city: str
        :param region: The name of the region
        :type region: str
        :return: The list of maps or None if no maps were found
        """
        with db_session() as session:
            if city is None:
                return session.query(Map).all()
            elif not region:
                return (
                    session.query(Map)
                    .join(Region, Region.id == Map.region_id)
                    .join(City, City.id == Region.city_id)
                    .filter(City.name.ilike(city))
                    .all()
                )
            else:
                return (
                    session.query(Map)
                    .join(Region, Region.id == Map.region_id)
                    .join(City, City.id == Region.city_id)
                    .filter(City.name.ilike(city))
                    .filter(Region.name.ilike(region))
                    .all()
                )

    @classmethod
    def get_matches_map(cls, address: str, city_name: str) -> Optional[Map]:
        """
        Gets the matches map for a given address and city

        :param address: The address
        :type address: str
        :param city_name: The city
        :type city_name: str
        :return: The map or None if no matches
        """

        with db_session() as session:
            return (
                session.query(Map)
                .join(Address, Address.map_id == Map.id)
                .join(Region, Region.id == Map.region_id)
                .join(City, City.id == Region.city_id)
                .filter(City.name.ilike(city_name))
                .filter(
                    (func.concat(Address.street, " ", Address.number).ilike(address))
                )
                .first()
            )

    @classmethod
    def _collect_maps_zip_xlsx(cls, zip_filepath, xlsx_filepath, city_name):
        """
        Collects all the maps from a zip file and extracts them

        :param zip_filepath: The path to the zip file
        :type zip_filepath: str
        :param xlsx_filepath: The path to the xlsx file
        :type xlsx_filepath: str
        :param city_name: The name of the city to extract
        :type city_name: str
        """

        def unpack_zip(filepath, city):
            """
            Unpacks the zip file into a dictionary

            :param filepath: The path to the zip file
            :type filepath: str
            :param city: The name of the city to extract
            :type city: str
            :return: A dictionary containing the extracted data
            """
            maps_struct = {city: {}}
            with ZipFile(filepath) as archive:
                for file in filter(
                    lambda filename: filename.endswith(".png"), archive.namelist()
                ):
                    region_name = Path(file).parent.name
                    if region_name not in maps_struct.get(city):
                        maps_struct[city][region_name] = {}
                    maps_struct[city][region_name][Path(file).stem] = {
                        "picture": archive.read(file),
                        "addresses": [],
                    }
            return maps_struct

        import locale

        locale.setlocale(locale.LC_ALL, ("ru_RU", "UTF-8"))

        maps = unpack_zip(zip_filepath, city_name)
        excel_table = load_workbook(xlsx_filepath)

        for city_regions in maps.values():
            for region, region_maps in city_regions.items():
                for map_name, map_values in region_maps.items():
                    map_addresses = list(
                        map(
                            lambda row: list(map(lambda cell: cell.value, row))[1:8],
                            filter(
                                lambda row: row[0].value and row[0].value == map_name,
                                excel_table[region].iter_rows(),
                            ),
                        )
                    )

                    map_values["addresses"] = []
                    for map_address in map_addresses:
                        if isinstance(map_address[0], datetime.datetime):
                            street: datetime.datetime = map_address[0]
                            str_street = f"{street.strftime('%d')} {street.strftime('%B').capitalize()}"
                            if str_street.startswith("0"):
                                str_street = str_street[1:]
                        elif isinstance(map_address[0], str):
                            str_street = map_address[0]
                        else:
                            print(map_address)
                            raise Exception

                        if isinstance(map_address[1], datetime.datetime):
                            number: datetime.datetime = map_address[1]
                            str_number = f"{number.day}/{number.month}"
                        elif isinstance(map_address[1], str):
                            str_number = map_address[1]
                        elif isinstance(map_address[1], float):
                            str_number = str(int(map_address[1]))
                        elif map_address[1] is None:
                            str_number = None
                        else:
                            print(map_address, map_name)
                            raise Exception

                        if isinstance(map_address[2], float):
                            int_floors = int(map_address[2])
                        elif isinstance(map_address[2], str):
                            if map_address[2].strip() == "-":
                                int_floors = None
                            elif map_address[2].isdigit():
                                int_floors = int(map_address[2])
                            else:
                                print(map_address, map_name)
                                raise Exception
                        elif isinstance(map_address[2], int):
                            int_floors = map_address[2]
                        else:
                            print(map_address, map_name)
                            raise Exception

                        if isinstance(map_address[3], float):
                            int_entrances = int(map_address[3])
                        elif isinstance(map_address[3], str):
                            if map_address[3].strip() in "-?":
                                int_entrances = None
                            elif map_address[3].isdigit():
                                int_entrances = int(map_address[3])
                            else:
                                print(map_address, map_name)
                                raise Exception
                        elif isinstance(map_address[3], int):
                            int_entrances = map_address[3]
                        else:
                            print(map_address, map_name)
                            raise Exception

                        if isinstance(map_address[4], float):
                            int_flats = int(map_address[4])
                        elif isinstance(map_address[4], str):
                            if map_address[4].strip() in "-?":
                                int_flats = None
                            elif map_address[4].isdigit():
                                int_flats = int(map_address[4])
                            else:
                                print(map_address, map_name)
                                raise Exception
                        elif isinstance(map_address[4], int):
                            int_flats = map_address[4]
                        else:
                            print(map_address, map_name)
                            raise Exception

                        str_comment = map_address[4]

                        map_values["addresses"].append(
                            [
                                str_street,
                                str_number,
                                int_floors,
                                int_entrances,
                                int_flats,
                                None,
                                str_comment,
                            ]
                        )
        return maps

    @classmethod
    def create_city(cls, city, zip_filepath, xlsx_filepath):
        """
        Create a new city file with the given data

        :param city: The city to create
        :type city: str
        :param zip_filepath: The path to the zip file
        :type zip_filepath: str
        :param xlsx_filepath: The path to the xlsx file
        :type xlsx_filepath: str
        """
        with db_session(mode="w") as session:
            maps = cls._collect_maps_zip_xlsx(zip_filepath, xlsx_filepath, city)

            for city, city_regions in maps.items():
                city_obj = City(city)
                session.add(city_obj)
                session.flush()
                session.refresh(city_obj)
                for city_region, region_maps in city_regions.items():
                    region_obj = Region(city_region, city_obj.id)
                    session.add(region_obj)
                    session.flush()
                    session.refresh(region_obj)
                    for region_map_number, region_map_vals in region_maps.items():
                        map_obj = Map(
                            region_map_number,
                            region_map_vals.get("picture"),
                            region_obj.id,
                        )
                        session.add(map_obj)
                        session.flush()
                        session.refresh(map_obj)
                        for address in region_map_vals.get("addresses"):
                            address_obj = Address(*address, map_obj.id)
                            session.add(address_obj)
                            session.flush()
                            session.refresh(address_obj)

    @classmethod
    def delete_city(cls, city_id):
        """
        Delete city from database with given id
        Cascade delete!

        :param city_id: Id of city in database
        :type city_id: int
        """

        with db_session(mode="w") as session:
            session.query(City).filter(City.id == city_id).delete()


# TODO async database
